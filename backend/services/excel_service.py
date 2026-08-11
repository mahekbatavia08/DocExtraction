"""
excel_service.py
────────────────
Standalone, Thread-Safe Excel Data Management & Export Service:
- Completely decoupled from OCR, Ollama, Qwen, and model router logic.
- Receives ONLY already-validated final JSON data and metadata.
- Storage Location: backend/data/excel/document_extraction.xlsx
- Workbook Structure:
    1. "Extraction Records" (Main Sheet - 15 Standard Columns)
    2. "Processing Summary" (Analytics Sheet)
- Features:
    - Thread-safe write lock (threading.Lock) to prevent file corruption
    - Configurable duplicate detection (DUPLICATE_MODE = "append" | "update")
    - Auto column width, freeze headers, auto-filter, and openpyxl formatting
    - Status reporting (Success, Partial, Failed)
"""

import os
import time
import io
import threading
from typing import Dict, Any, List, Optional
import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

EXCEL_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "excel")
EXCEL_FILE_PATH = os.path.join(EXCEL_DIR, "document_extraction.xlsx")

DUPLICATE_MODE = os.getenv("DUPLICATE_MODE", "append").lower()  # "append" or "update"

COLUMNS = [
    "Timestamp",
    "Document Type",
    "Name",
    "Father's Name",
    "Date of Birth",
    "PAN Number",
    "Address",
    "Pincode",
    "City",
    "State",
    "Confidence",
    "Model Used",
    "Fallback Used",
    "Processing Time",
    "Status"
]

class ExcelService:
    def __init__(self, file_path: str = EXCEL_FILE_PATH):
        self.file_path = file_path
        self.excel_dir = os.path.dirname(self.file_path)
        self.lock = threading.RLock()
        self._ensure_workbook_exists()

    def _ensure_workbook_exists(self):
        """Creates data directory and initializes Excel workbook if not present."""
        os.makedirs(self.excel_dir, exist_ok=True)

        with self.lock:
            if not os.path.exists(self.file_path):
                self._create_fresh_workbook()

    def _create_fresh_workbook(self):
        """Creates a brand new document_extraction.xlsx with structured sheets."""
        if OPENPYXL_AVAILABLE:
            wb = openpyxl.Workbook()
            # Main Sheet
            ws1 = wb.active
            ws1.title = "Extraction Records"
            ws1.append(COLUMNS)
            self._apply_header_styles(ws1)

            # Summary Sheet
            ws2 = wb.create_sheet(title="Processing Summary")
            ws2.append(["Metric Name", "Metric Value"])
            summary_headers = ws2[1]
            for cell in summary_headers:
                cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

            self._update_summary_sheet_data(ws1, ws2)
            wb.save(self.file_path)
        else:
            df_records = pd.DataFrame(columns=COLUMNS)
            df_records.to_csv(self.file_path, index=False)

    def _apply_header_styles(self, ws):
        """Applies openpyxl table styles: headers, borders, freeze panes, auto-filter."""
        if not OPENPYXL_AVAILABLE:
            return

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )

        for col_num, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        ws.freeze_panes = "A2"
        max_col_letter = get_column_letter(len(COLUMNS))
        ws.auto_filter.ref = f"A1:{max_col_letter}{ws.max_row}"

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    def record_extraction(self, final_data: Dict[str, Any], metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        Public entrypoint for inserting a final validated extraction result into Excel.
        Completely decoupled from AI pipeline — receives ONLY JSON + Metadata.
        Thread-safe execution.
        """
        try:
            with self.lock:
                self._ensure_workbook_exists()
                row_dict = self._build_row_dict(final_data, metadata)
                
                if OPENPYXL_AVAILABLE and os.path.exists(self.file_path):
                    wb = openpyxl.load_workbook(self.file_path)
                    ws1 = wb["Extraction Records"] if "Extraction Records" in wb.sheetnames else wb.active
                    ws2 = wb["Processing Summary"] if "Processing Summary" in wb.sheetnames else wb.create_sheet(title="Processing Summary")

                    # Handle duplicates
                    row_updated = False
                    if DUPLICATE_MODE == "update":
                        target_pan = row_dict["PAN Number"]
                        target_name = row_dict["Name"]
                        if target_pan and target_pan != "Not Found":
                            for r in range(2, ws1.max_row + 1):
                                pan_val = str(ws1.cell(row=r, column=6).value or "")
                                name_val = str(ws1.cell(row=r, column=3).value or "")
                                if pan_val == target_pan or (target_name and name_val == target_name):
                                    for c_idx, col_name in enumerate(COLUMNS, 1):
                                        ws1.cell(row=r, column=c_idx, value=row_dict[col_name])
                                    row_updated = True
                                    break

                    if not row_updated:
                        row_vals = [row_dict[col] for col in COLUMNS]
                        ws1.append(row_vals)

                    # Update formatting & summary sheet
                    self._apply_header_styles(ws1)
                    self._update_summary_sheet_data(ws1, ws2)
                    wb.save(self.file_path)
                else:
                    df = pd.DataFrame([row_dict])
                    file_exists = os.path.exists(self.file_path) and os.path.getsize(self.file_path) > 0
                    df.to_csv(self.file_path, mode='a' if file_exists else 'w', header=not file_exists, index=False)

                return {
                    "success": True,
                    "file_path": self.file_path,
                    "duplicate_mode": DUPLICATE_MODE,
                    "record_status": row_dict["Status"]
                }
        except Exception as e:
            # Excel failure MUST NOT crash the main AI extraction response
            return {
                "success": False,
                "error": f"Excel Write Warning: {str(e)}"
            }

    def _build_row_dict(self, final_data: Dict[str, Any], metadata: Dict[str, Any]) -> Dict[str, str]:
        """Maps final JSON + Metadata into standardized row dictionary."""
        def get_val(keys: List[str]) -> str:
            for k in keys:
                v = final_data.get(k)
                if v and str(v).upper() not in ["NOT FOUND", "N/A", "NULL", "NONE"]:
                    return str(v).strip()
            return "Not Found"

        doc_type = metadata.get("document_type") or final_data.get("document_type") or "General Document"
        conf_raw = metadata.get("confidence", 80.0)
        conf_str = f"{round(float(conf_raw), 1)}%" if isinstance(conf_raw, (int, float)) else str(conf_raw)
        
        proc_time = metadata.get("processing_time", 0.0)
        time_str = f"{round(float(proc_time), 2)}s" if isinstance(proc_time, (int, float)) else str(proc_time)

        # Status Flag Evaluation
        status = metadata.get("status", "")
        if not status:
            if float(conf_raw) >= 80.0 and not metadata.get("fallback_used"):
                status = "Success"
            elif metadata.get("fallback_used") or float(conf_raw) >= 60.0:
                status = "Partial"
            else:
                status = "Failed"
        elif status in ["success", "rule_fallback"]:
            status = "Success" if status == "success" else "Partial"

        return {
            "Timestamp": metadata.get("processing_timestamp") or time.strftime("%Y-%m-%d %H:%M:%S"),
            "Document Type": doc_type,
            "Name": get_val(["name", "Cardholder Name", "Name"]),
            "Father's Name": get_val(["father_name", "Father's Name", "Father Name"]),
            "Date of Birth": get_val(["dob", "Date of Birth", "DOB"]),
            "PAN Number": get_val(["pan_number", "PAN Number", "PAN"]),
            "Address": get_val(["address", "Address", "full_address"]),
            "Pincode": get_val(["pincode", "Pincode", "PIN Code"]),
            "City": get_val(["city", "City"]),
            "State": get_val(["state", "State"]),
            "Confidence": conf_str,
            "Model Used": metadata.get("model_used", "Rule-Based Engine"),
            "Fallback Used": "Yes" if metadata.get("fallback_used") else "No",
            "Processing Time": time_str,
            "Status": status
        }

    def _update_summary_sheet_data(self, ws1, ws2):
        """Recalculates analytics summary sheet metrics."""
        if not OPENPYXL_AVAILABLE:
            return

        total_records = ws1.max_row - 1
        if total_records <= 0:
            return

        success_cnt = 0
        partial_cnt = 0
        failed_cnt = 0
        fallback_cnt = 0
        model_counts: Dict[str, int] = {}
        total_conf = 0.0
        total_time = 0.0

        for r in range(2, ws1.max_row + 1):
            st = str(ws1.cell(row=r, column=15).value or "").strip()
            if st == "Success":
                success_cnt += 1
            elif st == "Partial":
                partial_cnt += 1
            else:
                failed_cnt += 1

            fb = str(ws1.cell(row=r, column=13).value or "").strip()
            if fb == "Yes":
                fallback_cnt += 1

            model = str(ws1.cell(row=r, column=12).value or "").strip()
            if model:
                model_counts[model] = model_counts.get(model, 0) + 1

            try:
                c_str = str(ws1.cell(row=r, column=11).value or "0").replace("%", "")
                total_conf += float(c_str)
            except ValueError:
                pass

            try:
                t_str = str(ws1.cell(row=r, column=14).value or "0").replace("s", "")
                total_time += float(t_str)
            except ValueError:
                pass

        avg_conf = round(total_conf / total_records, 1)
        avg_time = round(total_time / total_records, 2)
        fallback_pct = round((fallback_cnt / total_records) * 100, 1)

        ws2.delete_rows(2, ws2.max_row + 1)
        summary_rows = [
            ["Total Documents Processed", total_records],
            ["Successful Extractions", success_cnt],
            ["Partial Extractions", partial_cnt],
            ["Failed Extractions", failed_cnt],
            ["Average Confidence", f"{avg_conf}%"],
            ["Average Processing Time", f"{avg_time}s"],
            ["Fallback Usage Percentage", f"{fallback_pct}%"],
        ]

        for m_name, m_cnt in model_counts.items():
            summary_rows.append([f"Model Usage ({m_name})", m_cnt])

        for row in summary_rows:
            ws2.append(row)

    def get_records(self) -> List[Dict[str, Any]]:
        """Retrieves stored records from Excel sheet or CSV fallback."""
        with self.lock:
            if not os.path.exists(self.file_path):
                return []
            try:
                if OPENPYXL_AVAILABLE:
                    df = pd.read_excel(self.file_path, sheet_name="Extraction Records")
                else:
                    df = pd.read_csv(self.file_path)
                return df.fillna("Not Found").to_dict(orient="records")
            except Exception:
                try:
                    df = pd.read_csv(self.file_path)
                    return df.fillna("Not Found").to_dict(orient="records")
                except Exception:
                    return []

    def get_summary(self) -> List[Dict[str, Any]]:
        """Retrieves summary metrics from Excel summary sheet."""
        with self.lock:
            if not os.path.exists(self.file_path):
                return []
            try:
                if OPENPYXL_AVAILABLE:
                    df = pd.read_excel(self.file_path, sheet_name="Processing Summary")
                    return df.fillna("N/A").to_dict(orient="records")
                return []
            except Exception:
                return []

    def export_custom_buffer(self, documents: List[Dict[str, Any]]) -> bytes:
        """Exports custom document list into in-memory Excel buffer."""
        rows = []
        for doc in documents:
            fields = doc.get("fields") or doc.get("data") or {}
            metadata = doc.get("metadata") or {}
            rows.append(self._build_row_dict(fields, metadata))

        df = pd.DataFrame(rows)
        output = io.BytesIO()
        try:
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Extracted Documents")
        except Exception:
            csv_str = df.to_csv(index=False)
            output.write(csv_str.encode("utf-8"))

        output.seek(0)
        return output.getvalue()

excel_service = ExcelService()
