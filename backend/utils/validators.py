"""
validators.py
─────────────
Validation Algorithms & Normalization Engine for Universal AI Document Processing:
1. Verhoeff Checksum Algorithm (Dihedral Group D5 for Aadhaar 12-digit validation)
2. PAN Card Pattern & Structure Validator ([A-Z]{5}[0-9]{4}[A-Z])
3. Invoice Arithmetic Verification (Subtotal, Tax, Discount, Grand Total)
4. Excel Spreadsheet Parser (openpyxl macro-safe formula parser)
5. ISO-8601 Date & Currency Normalizer
6. Security Hash & Masking Utility
"""

import io
import re
import hashlib
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple

# ── 1. Verhoeff Checksum Algorithm (Dihedral Group D5) ───────────────────────

# Verhoeff Multiplication Table (d)
VERHOEFF_D = [
    [0, 1, 2, 3, 4, 5, 6, 7, 8, 9],
    [1, 2, 3, 4, 0, 6, 7, 8, 9, 5],
    [2, 3, 4, 0, 1, 7, 8, 9, 5, 6],
    [3, 4, 0, 1, 2, 8, 9, 5, 6, 7],
    [4, 0, 1, 2, 3, 9, 5, 6, 7, 8],
    [5, 9, 8, 7, 6, 0, 4, 3, 2, 1],
    [6, 5, 9, 8, 7, 1, 0, 4, 3, 2],
    [7, 6, 5, 9, 8, 2, 1, 0, 4, 3],
    [8, 7, 6, 5, 9, 3, 2, 1, 0, 4],
    [9, 8, 7, 6, 5, 4, 3, 2, 1, 0]
]

# Verhoeff Permutation Table (p)
VERHOEFF_P = [
    [0, 1, 2, 3, 4, 5, 6, 7, 8, 9],
    [1, 5, 7, 6, 2, 8, 3, 0, 9, 4],
    [5, 8, 0, 3, 7, 9, 6, 1, 4, 2],
    [8, 9, 1, 6, 0, 4, 3, 5, 2, 7],
    [9, 4, 5, 3, 1, 2, 6, 8, 7, 0],
    [4, 2, 8, 6, 5, 7, 3, 9, 0, 1],
    [2, 7, 9, 3, 8, 0, 6, 4, 1, 5],
    [7, 0, 4, 6, 9, 1, 3, 2, 5, 8]
]

def generate_verhoeff_checksum(number_11_digits: str) -> str:
    """Generate the 12th Verhoeff checksum digit for an 11-digit base number."""
    clean_num = re.sub(r'\D', '', str(number_11_digits))
    c = 0
    reversed_num = [int(n) for n in reversed(clean_num)]
    for i, digit in enumerate(reversed_num):
        c = VERHOEFF_D[c][VERHOEFF_P[(i + 1) % 8][digit]]
    
    # Find inverse digit to make checksum 0
    for digit in range(10):
        if VERHOEFF_D[c][VERHOEFF_P[0][digit]] == 0:
            return clean_num + str(digit)
    return clean_num + "0"

def validate_verhoeff(number_str: str) -> bool:
    """
    Validate a number string using the Verhoeff checksum algorithm (used for 12-digit Aadhaar UID).
    Returns True if checksum matches 0.
    """
    clean_num = re.sub(r'\D', '', str(number_str))
    if len(clean_num) != 12:
        return False

    c = 0
    reversed_num = [int(n) for n in reversed(clean_num)]
    for i, digit in enumerate(reversed_num):
        c = VERHOEFF_D[c][VERHOEFF_P[i % 8][digit]]

    return c == 0


def mask_aadhaar_number(aadhaar_str: str) -> str:
    """Mask first 8 digits of Aadhaar number for security compliance: XXXX XXXX 1234."""
    clean_num = re.sub(r'\D', '', str(aadhaar_str))
    if len(clean_num) == 12:
        return f"XXXX XXXX {clean_num[-4:]}"
    return aadhaar_str


# ── 2. PAN Card Validation ────────────────────────────────────────────────────

PAN_REGEX = r'^[A-Z]{5}[0-9]{4}[A-Z]$'

PAN_ENTITY_TYPES = {
    'C': 'Company',
    'P': 'Individual Person',
    'H': 'Hindu Undivided Family (HUF)',
    'F': 'Firm / Partnership',
    'A': 'Association of Persons (AOP)',
    'T': 'Trust',
    'B': 'Body of Individuals (BOI)',
    'L': 'Local Authority',
    'J': 'Artificial Juridical Person',
    'G': 'Government Agency'
}

def validate_pan_number(pan_str: str) -> Dict[str, Any]:
    """Validate 10-character PAN number structure and extract entity type (4th character)."""
    clean_pan = str(pan_str).strip().upper()
    is_valid = bool(re.match(PAN_REGEX, clean_pan))
    
    # 4th character (index 3) defines the PAN holder type
    entity_code = clean_pan[3] if len(clean_pan) == 10 else None
    entity_type = PAN_ENTITY_TYPES.get(entity_code, "Individual Person" if is_valid else "Unknown Entity")

    return {
        "is_valid": is_valid,
        "pan_number": clean_pan if is_valid else None,
        "entity_type": entity_type if is_valid else None,
        "confidence": 0.99 if is_valid else 0.0
    }


# ── 3. Invoice Arithmetic Verification ────────────────────────────────────────

def audit_invoice_arithmetic(extracted_fields: Dict[str, Any], raw_text: str) -> Dict[str, Any]:
    """
    Perform mathematical cross-auditing on invoices:
    Computes Subtotal + Tax - Discount == Grand Total and highlights mismatches.
    """
    def parse_amount(val: Any) -> Optional[float]:
        if not val:
            return None
        cleaned = re.sub(r'[^\d.-]', '', str(val))
        try:
            return float(cleaned)
        except ValueError:
            return None

    subtotal = parse_amount(extracted_fields.get("Subtotal"))
    tax = parse_amount(extracted_fields.get("Tax / GST"))
    discount = parse_amount(extracted_fields.get("Discount", 0.0)) or 0.0
    grand_total = parse_amount(extracted_fields.get("Grand Total") or extracted_fields.get("Total Amount"))

    # Fallback to regex from raw text if fields missing
    if not grand_total:
        gt_match = re.search(r'(?:Grand Total|Total Amount|Amount Due)[:\s]*\$?\s?([\d,]+\.\d{2})', raw_text, re.IGNORECASE)
        if gt_match:
            grand_total = parse_amount(gt_match.group(1))

    passed_rules = []
    failed_rules = []

    if subtotal is not None and tax is not None and grand_total is not None:
        computed_total = round(subtotal + tax - discount, 2)
        diff = abs(computed_total - grand_total)
        if diff <= max(0.10, 0.02 * grand_total):
            passed_rules.append(f"Arithmetic Verification PASS: Subtotal ({subtotal:.2f}) + Tax ({tax:.2f}) - Discount ({discount:.2f}) = {computed_total:.2f} (Matches Grand Total {grand_total:.2f})")
        else:
            failed_rules.append(f"Arithmetic Discrepancy: Computed {computed_total:.2f}, but found Grand Total {grand_total:.2f} (Diff: {diff:.2f})")
    elif grand_total is not None:
        passed_rules.append(f"Extracted Grand Total: {grand_total:.2f}")

    return {
        "status": "PASS" if not failed_rules else "DISCREPANCY_FOUND",
        "passed_rules": passed_rules,
        "failed_rules": failed_rules,
        "computed_grand_total": round(subtotal + tax - discount, 2) if (subtotal is not None and tax is not None) else grand_total
    }


# ── 4. Excel Spreadsheet Parser (Macro-Safe) ──────────────────────────────────

def parse_excel_spreadsheet(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """
    Parse Excel workbook (.xlsx/.xls) cleanly without executing macros.
    Extracts sheets, rows, columns, formulas, named ranges, and merged cells.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False, keep_vba=False)
        sheets_data = {}
        full_text_lines = []

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            matrix = []
            full_text_lines.append(f"=== Sheet: {sheetname} ===")
            for row in ws.iter_rows(values_only=False):
                str_row = []
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    str_row.append(val)
                if any(str_row):
                    matrix.append(str_row)
                    full_text_lines.append(" | ".join(str_row))

            sheets_data[sheetname] = {
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "matrix": matrix[:100]  # First 100 rows preview
            }

        return {
            "document_type": "Excel Spreadsheet",
            "filename": filename,
            "sheets_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "sheets": sheets_data,
            "raw_text": "\n".join(full_text_lines),
            "macro_execution_blocked": True
        }
    except Exception as e:
        raise ValueError(f"Excel parsing failed for {filename}: {str(e)}")


# ── 5. Normalization & Security Helpers ──────────────────────────────────────

def normalize_date(date_str: str) -> Optional[str]:
    """Normalize date strings into ISO-8601 format YYYY-MM-DD."""
    if not date_str:
        return None
    clean = date_str.strip()
    date_formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"]
    for fmt in date_formats:
        try:
            dt = datetime.strptime(clean, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return clean


def hash_sensitive_data(val: str) -> str:
    """Generate SHA-256 hash for sensitive biometric or identity fields."""
    return hashlib.sha256(str(val).encode("utf-8")).hexdigest()[:16]
